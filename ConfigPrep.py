import datetime
import os
import os.path
from os import path
import re
import pyinputplus as pyip
from openpyxl import load_workbook
from CUCM_Groups_Detector import FindCUCM_Group
from ISDN_Detector import FindISDN
import warnings
warnings.filterwarnings('ignore')
#ttt
SiteCode = pyip.inputStr(prompt='Enter a Site Code: ')
if len(SiteCode) != 5:
    print("Not a valid site code.Example - EDx01. You have in your input",len(SiteCode) )
    exit()
#SiteCode = "EGx01"
py_file_location=os.path.dirname(os.path.abspath(__file__))
os.chdir(py_file_location)
print("Plase review:\n","Site Code:",SiteCode,"Script and DCS locatin folder:",py_file_location,"Config files location folder: /",SiteCode )
#print ("directory exists:" + str(path.exists(SiteCode)))
if path.exists(SiteCode) != True:
    os.makedirs(SiteCode)
os.chdir(SiteCode)

rootdir = os.getcwd()

regex = re.compile(r"(?=.*DataCollection).*(?=.*" + str(SiteCode) + ").*(?=.*xlsm$)",re.IGNORECASE)
#regex = re.compile('(?=.*DataCollection).*(?=.*xlsm$)')
#regex = re.compile("(?=.*DataCollection).*" + str((SiteCode, re.IGNORECASE))+".*(?=.*xlsm$)")
DCS_file = "Default"
for root, dirs, files in os.walk(py_file_location):
  for file in files:
 #   print("List of files in directroy",file)
    if regex.match(file):
        print("Found file",file, "in ",root)
        DCS_file=file
if DCS_file == "Default":
    print("Missing DCS file with SiteCode",SiteCode,"in direcory",py_file_location)
    exit()
#Second file validation
#if DCS_file == "Default":
#    print("Missing DCS file in direcory",os.getcwd())
#    exit()
#print("Second file validation. Matching",SiteCode, "with",DCS_file)
re_pattern = re.compile(SiteCode, re.IGNORECASE)
#match2 = re_pattern.findall(DCS_file)
#print(match2)
#if str(match2) == str(""):
#    print("Missing DCS file with SiteCode",SiteCode,"in direcory",py_file_location)
#    exit()
#elif SiteCode.lower() == match2[0].lower():
#    print("Site Code is",SiteCode, "DCS file detected with",match2)
#else:
#    print("Missing DCS file in direcory",os.getcwd())
#    exit()
#####
#Excel read data#
#xl = pd.ExcelFile(DCS_file)
# Print the sheet names
#print(xl.sheet_names)
#df1 = xl.parse('Allgemeines')
#print(df1)
wb = load_workbook(py_file_location+"\\"+DCS_file)
# Get sheet names
#print(wb.sheetnames)
sheet = wb["Allgemeines"]
sheet2 = wb["Userdaten"]
#print("PSTN GW hostname is",sheet["C31"].value)
#print("Analog GW hostname is",sheet["D31"].value)
#print("Site",SiteCode,"phone format nummer is:",sheet["C15"].value,sheet["C16"].value)
#Check and validation
## Checking the site extension len
Ext = str(sheet2["L3"].value)
#print("Extension len is",len(Ext))
if len(Ext) == 3:
    PSTN_Ext = "..."
    PSTN_ExtX = "XXX"
elif len(Ext) == 2:
    PSTN_Ext = ".."
    PSTN_ExtX = "XX"
else:
    print("Error, something is not right with the Site extension numbers in sheet", sheet2)
#
match2 = re_pattern.findall(sheet["C31"].value)
if SiteCode.lower() == match2[0].lower():
    print("Detected",match2, "PSTN in hostname",sheet["C31"].value)
else:
    print("error, no match betwwen",SiteCode,"and",str(match2[0]))
    exit()
match2 = re_pattern.findall(sheet["D31"].value)
if SiteCode.lower() == match2[0].lower():
    print("Detected",match2, "in Analog hostname",sheet["D31"].value)
else:
    print("error")
    exit()
re_pattern2 = re.compile("\+49")
match2 = re_pattern2.findall(sheet["C15"].value)
if match2[0]=="+49":
    print("Detected",match2, "in phone number ",sheet["C15"].value,sheet["C16"].value,PSTN_Ext)
else:
    print("error, not a valid phone number")
    exit()

##########################################################################################
SiteDataDB = [SiteCode,sheet["C7"].value]
print("Detected",str(SiteDataDB))
DailPlanDB = ([int(sheet["C15"].value),sheet["C16"].value,PSTN_Ext,PSTN_ExtX])
PSTN_GW_DB = [sheet["C31"].value,sheet["C33"].value]
PSTNno49 = str([DailPlanDB[0]])
PSTNno49_REGEX = re.sub(r"\D", "", PSTNno49, flags=re.I)
AnalogGW1_DB = [sheet["D31"].value,sheet["D33"].value]
AnalogGW2_DB = [sheet["E31"].value,sheet["E33"].value]
DNS_DB = [sheet["C48"].value,sheet["C49"].value]
NTP_DB = [sheet["C45"].value,sheet["C46"].value]
#removeFX = [re.sub('[^a-zA-Z0-9]+', '', _) for _ in DailPlanDB]
#print(removeFX)
##########################################################################################
#Find the CUCM Group for the site. Important
SearchCUCM = FindCUCM_Group(sheet["C7"].value)
if SearchCUCM == "NotFound":
    print("From the DCS file, we cannot find the CUCM Group for site:",SiteDataDB[1]+". This will require manual selection")
    CUCM_Group = pyip.inputMenu(['CMGroup1_2','CMGroup2_1','CMGroup4_5','CMGroup5_4','CMGroup3_6','CMGroup6_3'], numbered=True)
    print("CallManager Group will be set manualy to: ",CUCM_Group)
else:
    CUCM_Group = SearchCUCM
    print("Detected (automatic) CallManager Group is: " + CUCM_Group[0])
##########################################################################################
####################################ISDN Detector ###############################################
#Find the ISDN type and count for the site. Important
SearchISDN = FindISDN(SiteCode,PSTN_GW_DB)
#if SearchCUCM == "NotFound":
#    CUCM_Group = pyip.inputMenu(['CMGroup1_2','CMGroup2_1','CMGroup4_5','CMGroup5_4','CMGroup3_6','CMGroup6_3'], numbered=True)
#    print("CallManager Group will be set manualy to: ",CUCM_Group)
#else:
#    CUCM_Group = SearchCUCM
#    print("Detected (automatic) CallManager Group is: " + CUCM_Group[0])
print("ISDN type is:",SearchISDN)
##########################################################################################
#print(PSTNno49_REGEX[2:])
print("#################################################################")
print("Site information:",SiteDataDB,CUCM_Group[0])
print("DialPlan:",DailPlanDB)
print("PSTN GW:",PSTN_GW_DB)
print("Analog GW-1:",AnalogGW1_DB)
print("Analog GW-2:",AnalogGW2_DB)
print("DNS Servers:",DNS_DB)
print("NTP Servers:",NTP_DB)
print("DCS file:",DCS_file)
print("#################################################################")
response = pyip.inputYesNo(prompt="is the information valid ? yes/no: ")
if response == "no":
    exit()
###################### Create Config File ########################################################
ConfigFile= open(str(SiteCode) + "_PSTN_config.txt","w+")
ConfigFile.write("########### Config File for PSTN Router " + str(PSTN_GW_DB[0]) + " ########### \r\n")
ConfigFile.close()
ConfigFile= open(str(SiteCode) + "_PSTN_config.txt","a+")
###################### Certificates #########################################################
ConfigFile.write("####### Certificates #######\n\r")
ConfigFile.write("crypto pki trustpoint CAPF-1cf4174d\n")
ConfigFile.write(" enrollment terminal\n")
ConfigFile.write(" revocation-check none\n\r")
ConfigFile.write("crypto pki authenticate CAPF-1cf4174d\n\r")
ConfigFile.write("-----BEGIN CERTIFICATE-----\n")
ConfigFile.write("MIIDwzCCAqugAwIBAgIQaBxF0gOptqqyr51Exc5iMDANBgkqhkiG9w0BAQsFADBx\n")
ConfigFile.write("MQswCQYDVQQGEwJERTEZMBcGA1UECgwQRnJlaXN0YWF0IEJheWVybjENMAsGA1UE\n")
ConfigFile.write("CwwEbGRidjEWMBQGA1UEAwwNQ0FQRi0xY2Y0MTc0ZDEPMA0GA1UECAwGQmF5ZXJu\n")
ConfigFile.write("MQ8wDQYDVQQHDAZCYXllcm4wHhcNMTkwNTA3MTAxODM1WhcNMjQwNTA1MTAxODM0\n")
ConfigFile.write("WjBxMQswCQYDVQQGEwJERTEZMBcGA1UECgwQRnJlaXN0YWF0IEJheWVybjENMAsG\n")
ConfigFile.write("A1UECwwEbGRidjEWMBQGA1UEAwwNQ0FQRi0xY2Y0MTc0ZDEPMA0GA1UECAwGQmF5\n")
ConfigFile.write("ZXJuMQ8wDQYDVQQHDAZCYXllcm4wggEiMA0GCSqGSIb3DQEBAQUAA4IBDwAwggEK\n")
ConfigFile.write("AoIBAQCTju/3D5G2kaWaJh9rquUuIfbq1mNf+954eaumKcEDPDOe04JbNGO+152A\n")
ConfigFile.write("NjcG6W3T7w2kI37E5UiYESHNyBqBPBZs9xjoakKYVIaCb/Xgeuiw354N+ym7e8zG\n")
ConfigFile.write("1uFCwYvzgNmfUyoWxbUuVzYlbJdIr93jmCQMsCGgDtd+dP5Rcm4A/iVyvIWzhIYz\n")
ConfigFile.write("EiJj3Ygi+QkqrfMMfe4neCTeIiWD4u5FExHsex1TM7lSdWqZzPDSz7zNLNQyDokt\n")
ConfigFile.write("TxlGUP52EJJj2XBUNqlO0CrCH/Rv/QCa8p/fjIT/wds2jh2oLTe9ri/O6lOznNDZ\n")
ConfigFile.write("Cb04yKNgdTrBStnbxL0V9ODz5KbXAgMBAAGjVzBVMAsGA1UdDwQEAwICpDATBgNV\n")
ConfigFile.write("HSUEDDAKBggrBgEFBQcDATAdBgNVHQ4EFgQUlyu1pKJY/2iRG3+jtXdU4ezsulIw\n")
ConfigFile.write("EgYDVR0TAQH/BAgwBgEB/wIBADANBgkqhkiG9w0BAQsFAAOCAQEAhbfOJJPIM+7d\n")
ConfigFile.write("dX1Tx9lKsOdDrWqoz0p1u3X4gxlOxxAhvJZRvOQnZI4yRah+Y0rCIPJwYe/XS0dc\n")
ConfigFile.write("1KknPVkus2FUwNYuGfvywDrAek5Dm9I2zXWpPiKjd/DtfbBnmGbM3aw8BmHTf4gy\n")
ConfigFile.write("ssyEAD29RmAGyeAbjwuqJxgpStQI3P6WddtpvIZK4dofblzaSh5lw9zdHlgIbmDG\n")
ConfigFile.write("KfNPkv7NOeRnULMnp8TYFXN/QytFvuPyQbPXszbwkW5s5pjHKqsjZU2nVWmpqpsJ\n")
ConfigFile.write("B4QGnzX3l4x1VnHOu1e6axsyQSYOjjm5PQ2rZ/zZqn5IDHktKvWH9A0SunrvGC85\n")
ConfigFile.write("lRsCurwLKA==\n")
ConfigFile.write("-----END CERTIFICATE-----\n\r")

ConfigFile.write("crypto pki trustpoint CAPF-d500f2a9\n")
ConfigFile.write(" enrollment terminal\n")
ConfigFile.write(" revocation-check none\n\r")
ConfigFile.write("crypto pki authenticate CAPF-d500f2a9\n\r")
ConfigFile.write("-----BEGIN CERTIFICATE-----\n")
ConfigFile.write("MIIDwzCCAqugAwIBAgIQfIAfE9XFp/PXePAEm2LKVzANBgkqhkiG9w0BAQsFADBx\n")
ConfigFile.write("MQswCQYDVQQGEwJERTEZMBcGA1UECgwQRnJlaXN0YWF0IEJheWVybjENMAsGA1UE\n")
ConfigFile.write("CwwEbGRidjEWMBQGA1UEAwwNQ0FQRi1kNTAwZjJhOTEPMA0GA1UECAwGQmF5ZXJu\n")
ConfigFile.write("MQ8wDQYDVQQHDAZCYXllcm4wHhcNMTkwMzI5MTYxMDUyWhcNMjQwMzI3MTYxMDUx\n")
ConfigFile.write("WjBxMQswCQYDVQQGEwJERTEZMBcGA1UECgwQRnJlaXN0YWF0IEJheWVybjENMAsG\n")
ConfigFile.write("A1UECwwEbGRidjEWMBQGA1UEAwwNQ0FQRi1kNTAwZjJhOTEPMA0GA1UECAwGQmF5\n")
ConfigFile.write("ZXJuMQ8wDQYDVQQHDAZCYXllcm4wggEiMA0GCSqGSIb3DQEBAQUAA4IBDwAwggEK\n")
ConfigFile.write("AoIBAQCvMKuuIdCAdb0PRPhEUyDpeNjCPLMVxpb/plvZpKqIpvqlQEriFbnvVkK7\n")
ConfigFile.write("ukrof3iYex1f6tU/y/u1RyJ9F6avD7FMsd8ggccx6vznl5EN5CUqcV3kBCmxF9RD\n")
ConfigFile.write("TS3hNuD6+R0VFiqkA7ulnpu/HT+mTu0rgbzO9n4CkhBu9mhX+nqubmZs/HSSE2r1\n")
ConfigFile.write("FPSBJCin3VDKP5JTROaSx9BXbNiHpIJM7KbjwE+xzJ6RZEO91Fs8CJA9Cl+KFxXN\n")
ConfigFile.write("3Z4CKfDUM4rOW0T56URDECGEpNhkz/PNYI07W8ggOQNURJiVmZyneql1jcQQ5uN7\n")
ConfigFile.write("PY4h3V9dY1nKaeATiTWeOtwa+3uJAgMBAAGjVzBVMAsGA1UdDwQEAwICpDATBgNV\n")
ConfigFile.write("HSUEDDAKBggrBgEFBQcDATAdBgNVHQ4EFgQU5jJ/c/Ekbl40GAnoj4+d/AAAIWsw\n")
ConfigFile.write("EgYDVR0TAQH/BAgwBgEB/wIBADANBgkqhkiG9w0BAQsFAAOCAQEAWgqO3SjFrXdU\n")
ConfigFile.write("tZkPQk6QgKAtRbMPVXYm/rkSeQzidvoIs5MiGvF5p4BjoiYrNwZ/3GHvjCODzZcU\n")
ConfigFile.write("jEIvkDuuR8dujR9j1OiTqXO/48x7vPxU3OB1xL09g5wG7l0vY6tYDsHKjIfcKWJQ\n")
ConfigFile.write("dm6EKRCKrv6Vsd7gxvv1yPfPHvbFjWOHbN1zNByG+lSUhYPHPn5Ypv2hcQFyotAZ\n")
ConfigFile.write("w6/njAaixmz5aL4RZAb6EJzx/D+uwhjMGDP7mkZptNARbUzLMuLyd4A7gNZw1KBn\n")
ConfigFile.write("eBoVyfuYgh1u0uU8w0zo/aVpRA45TPuVMOmhCnyucecwse1WHtefBq07kgLRj0kH\n")
ConfigFile.write("IGED67G1ug==\n")
ConfigFile.write("-----END CERTIFICATE-----\n\r")

ConfigFile.write("crypto key generate rsa general-keys label " + str(PSTN_GW_DB[0]) + " modulus 2048\n\r")

ConfigFile.write("crypto pki trustpoint " + str(PSTN_GW_DB[0]) + " \n")
ConfigFile.write(" enrollment terminal\n")
ConfigFile.write(" serial-number none\n")
ConfigFile.write(" fqdn none\n")
ConfigFile.write(" ip-address none\n")
ConfigFile.write(" subject-name CN=" + str(PSTN_GW_DB[0]) + ".juwin.bayern.de\n")
ConfigFile.write(" revocation-check none\n")
ConfigFile.write(" rsakeypair "  + str(PSTN_GW_DB[0]) + " 2048 2048\n")
ConfigFile.write(" hash sha256\n\r")

ConfigFile.write("crypto pki enroll " + str(PSTN_GW_DB[0]) + "\n\r")

ConfigFile.write("####### Install first the ROOT cert #######\n")
ConfigFile.write("crypto pki authenticate " + str(PSTN_GW_DB[0]) + "\n\r") 

ConfigFile.write("####### ROOT Infrastructure #######\n\r")

ConfigFile.write("-----BEGIN CERTIFICATE-----\n")
ConfigFile.write("MIIGzTCCBLWgAwIBAgITTQAAAAL73ahB3zfUAAAAAAAAAjANBgkqhkiG9w0BAQ0F\n")
ConfigFile.write("ADAbMRkwFwYDVQQDExBKdXN0aXotUm9vdC1DQS0xMB4XDTE3MDcyNzEzMTA0OFoX\n")
ConfigFile.write("DTMyMDcyNzEzMjA0OFowZzESMBAGCgmSJomT8ixkARkWAmRlMRYwFAYKCZImiZPy\n")
ConfigFile.write("LGQBGRYGYmF5ZXJuMRUwEwYKCZImiZPyLGQBGRYFanV3aW4xIjAgBgNVBAMMGUp1\n")
ConfigFile.write("c3Rpei1JbmZyYXN0cnVrdHVyLUNBLTEwggEiMA0GCSqGSIb3DQEBAQUAA4IBDwAw\n")
ConfigFile.write("ggEKAoIBAQDquommbj+zoM2OBFmWJSRxKSupMWgH1Ops4jhK09fkxtJjLaEdp3m/\n")
ConfigFile.write("qjDW2ZZ5ntPHBWw5REE4EozjFmF2dtS9eFx9ll6+SsJw6oJvU0f3CHq/4DZNQQ/a\n")
ConfigFile.write("UxSlFR/huCkbmWXEyjjZ2miqYuc9HXHrE81g1FYx/Pn0IgU0e8Ry3e+GbRfQkGv0\n")
ConfigFile.write("CWrIBLMBz6M3wRyuHtwIYcCNhOPCLLFzZOIGFpkT/EyJN4dZXg1AiSgOIVy/R2L/\n")
ConfigFile.write("Z7+xYk8elEAthql1Mmu44XTZWF61+Hc2rm6j84ULoJWszG09pFufMeTNFs6+9GgQ\n")
ConfigFile.write("wPZGQZj++xSQa3Ka8usVLdjFB5JEyi/lAgMBAAGjggK8MIICuDAOBgNVHQ8BAf8E\n")
ConfigFile.write("BAMCAQYwEAYJKwYBBAGCNxUBBAMCAQAwHQYDVR0OBBYEFGR+AqEO/0NAI/0DRakZ\n")
ConfigFile.write("hUGyAxEFMBkGCSsGAQQBgjcUAgQMHgoAUwB1AGIAQwBBMA8GA1UdEwEB/wQFMAMB\n")
ConfigFile.write("Af8wHwYDVR0jBBgwFoAU+ASb1NP/L1CtiDd8LzjABihkUkswggENBgNVHR8EggEE\n")
ConfigFile.write("MIIBADCB/aCB+qCB94aBw2xkYXA6Ly8vQ049SnVzdGl6LVJvb3QtQ0EtMSxDTj1O\n")
ConfigFile.write("eHgwM1pTSDAyMXcsQ049Q0RQLENOPVB1YmxpYyUyMEtleSUyMFNlcnZpY2VzLENO\n")
ConfigFile.write("PVNlcnZpY2VzLENOPUNvbmZpZ3VyYXRpb24sREM9anV3aW4sREM9YmF5ZXJuLERD\n")
ConfigFile.write("PWRlP2NlcnRpZmljYXRlUmV2b2NhdGlvbkxpc3Q/YmFzZT9vYmplY3RDbGFzcz1j\n")
ConfigFile.write("UkxEaXN0cmlidXRpb25Qb2ludIYvaHR0cDovL2NybC5qdXdpbi5iYXllcm4uZGUv\n")
ConfigFile.write("SnVzdGl6LVJvb3QtQ0EtMS5jcmwwggEVBggrBgEFBQcBAQSCAQcwggEDMIG2Bggr\n")
ConfigFile.write("BgEFBQcwAoaBqWxkYXA6Ly8vQ049SnVzdGl6LVJvb3QtQ0EtMSxDTj1BSUEsQ049\n")
ConfigFile.write("UHVibGljJTIwS2V5JTIwU2VydmljZXMsQ049U2VydmljZXMsQ049Q29uZmlndXJh\n")
ConfigFile.write("dGlvbixEQz1qdXdpbixEQz1iYXllcm4sREM9ZGU/Y0FDZXJ0aWZpY2F0ZT9iYXNl\n")
ConfigFile.write("P29iamVjdENsYXNzPWNlcnRpZmljYXRpb25BdXRob3JpdHkwSAYIKwYBBQUHMAKG\n")
ConfigFile.write("PGh0dHA6Ly9jcmwuanV3aW4uYmF5ZXJuLmRlL054eDAzWlNIMDIxd19KdXN0aXot\n")
ConfigFile.write("Um9vdC1DQS0xLmNydDANBgkqhkiG9w0BAQ0FAAOCAgEAkCtlyXAYcjuP9CnWsuMD\n")
ConfigFile.write("AOP1To884O+zbAXSRsan6N/2HS+kyCl1O7YL/69nlV+e4T2ElcP9p7lf41nrrTxI\n")
ConfigFile.write("hklU8tmKbG6F6S6VATkrj0NgbIg9KZ6bH7rHgf5soVvjbOtK8/nx8WXqujuBPyRk\n")
ConfigFile.write("iOt6Aax+Xbs5i0KO35shOAKFtrwelZh/zE2DPOvEaip+ZG2ZbA6BzXUc+5i+IYL8\n")
ConfigFile.write("Of0Z1O1/q7meiftz6xZ4C9JPRRynyHvvD1jaE4gkeFZ3ExHQLNHWFPPmSrQwUPp2\n")
ConfigFile.write("Yt560ARe8xGZlbLsTXKf2SAZi9smw6zlsHRYCmTX6OEdX+IP6tz61t/8IGAcIR7P\n")
ConfigFile.write("PNr3yCG8BQOzx4341oQJBRXOCvKwyey1dmSfdUjD50DdHHin52W8tDu46VTF7wwg\n")
ConfigFile.write("g0vv9moYlIsY2Fi37hQVHs4kizFBtuFBrRAP1U6SBnMDmifIWSAy3mVljosZfmU+\n")
ConfigFile.write("Y6pPmP2a8FoRHp5e8EXf5LWDUFMka7+C9na5IVkGt/yqoI/DWLfyApKRLnF5YNJu\n")
ConfigFile.write("XNH//pjImyDl532rYmM3QCQrka8E0EEllvEtb+axkZ8Hu1CaYNeo6y+PZVlmi7Tp\n")
ConfigFile.write("UOqpjXBo7zV3a5LMyfABuEBZEG+xw5xcZ2x8ZaVUA+Ho+jbSUFJ9g1JHr6l1beuy\n")
ConfigFile.write("qSNFLBO103r4zr6+t8n+gno=\n")
ConfigFile.write("-----END CERTIFICATE-----\n\r")

ConfigFile.write("####### Install the Justit CA signed certificate #######\n")
ConfigFile.write("crypto pki import " + str(PSTN_GW_DB[0]) + " certificate \n\r")
########################################################################################
############################## Prime SNMP ##############################################
ConfigFile.write("####### Prime SNMP #######\n\r")
ConfigFile.write("snmp-server view UC_SNMP dod included\n")
ConfigFile.write("snmp-server host 10.216.3.17 version 3 auth UC_SNMP\n")
ConfigFile.write("snmp-server group UC_SNMP v3 auth read UC_SNMP write UC_SNMP\n")
ConfigFile.write("snmp-server user xx UC_SNMP v3 auth sha xx!\n\r")
########################################################################################
############################## VoIP static #############################################
ConfigFile.write("####### Static VoIP #######\n\r")
ConfigFile.write("voice call send-alert\n")
ConfigFile.write("voice rtp send-recv\n")
ConfigFile.write("stcapp security tls-version v1.2\n")
ConfigFile.write("trunk group PSTN\n")
ConfigFile.write("voice service voip\n")
ConfigFile.write(" ip address trusted list\n")
ConfigFile.write(" ipv4 10.216.1.6\n")
ConfigFile.write(" ipv4 10.216.3.10\n")
ConfigFile.write(" ipv4 10.216.3.14\n")
ConfigFile.write(" ipv4 10.216.3.15\n")
ConfigFile.write(" ipv4 10.216.1.7\n")
ConfigFile.write(" ipv4 10.216.1.10\n")
ConfigFile.write(" ipv4 10.216.1.11\n")
ConfigFile.write(" ipv4 10.217.122.10\n")
ConfigFile.write("srtp fallback\n")
ConfigFile.write(" allow-connections sip to sip\n")
ConfigFile.write(" no supplementary-service sip moved-temporarily\n")
ConfigFile.write(" no supplementary-service sip refer\n")
ConfigFile.write(" supplementary-service media-renegotiate\n")
ConfigFile.write(" fax protocol none\n")
ConfigFile.write(" modem passthrough nse codec g711alaw\n")
ConfigFile.write(" sip\n")
ConfigFile.write("  bind control source-interface Loopback0\n")
ConfigFile.write("  bind media source-interface Loopback0\n")
ConfigFile.write("  session transport tcp tls\n")
ConfigFile.write("  registrar server expires max 600 min 60\n")
ConfigFile.write("  srtp negotiate cisco\n")
ConfigFile.write("  early-offer forced\n")
ConfigFile.write("  midcall-signaling passthru\n\r")

ConfigFile.write("voice-card 0/4\n")
ConfigFile.write(" dsp services dspfarm\n")
ConfigFile.write(" no watchdog\n\r")

ConfigFile.write("call-manager-fallback\n")
ConfigFile.write(" transport-tcp-tls v1.2\n")
ConfigFile.write(" max-conferences 8 gain -6\n")
ConfigFile.write(" transfer-system full-consult\n")
ConfigFile.write(" timeouts interdigit 5\n")
ConfigFile.write(" ip source-address " + str(PSTN_GW_DB[1]) + " port 2000\n")
ConfigFile.write(" max-ephones 100\n")
ConfigFile.write(" max-dn 200\n")
ConfigFile.write(" system message primary Lokale Notfalltelefonie\n")
ConfigFile.write(" transfer-pattern 1.\n")
ConfigFile.write(" transfer-pattern 2.\n")
ConfigFile.write(" transfer-pattern 3.\n")
ConfigFile.write(" transfer-pattern 5.\n")
ConfigFile.write(" transfer-pattern 6.\n")
ConfigFile.write(" transfer-pattern 7.\n")
ConfigFile.write(" transfer-pattern 8.\n")
ConfigFile.write(" transfer-pattern 9.\n")
ConfigFile.write(" transfer-pattern T\n")
ConfigFile.write(" translation-profile incoming SRST-INCOMING\n")
ConfigFile.write(" call-forward pattern 1.\n")
ConfigFile.write(" call-forward pattern 2.\n")
ConfigFile.write(" call-forward pattern 3.\n")
ConfigFile.write(" call-forward pattern 5.\n")
ConfigFile.write(" call-forward pattern 6.\n")
ConfigFile.write(" call-forward pattern 7.\n")
ConfigFile.write(" call-forward pattern 8.\n")
ConfigFile.write(" call-forward pattern 9.\n")
ConfigFile.write(" moh enable-g711 flash:SampleAudioSource.alaw.wav\n")
ConfigFile.write(" moh g729 flash:SampleAudioSource.g729.wav\n")
ConfigFile.write(" time-format 24\n")
ConfigFile.write(" date-format dd-mm-yy\n\r")

ConfigFile.write("no credentials\n")
ConfigFile.write("credentials\n")
ConfigFile.write(" ip source-address "  + str(PSTN_GW_DB[1]) + " port 2445\n")
ConfigFile.write(" trustpoint " + str(PSTN_GW_DB[0]) + " \r\n")

ConfigFile.write("sip-ua\n")
ConfigFile.write(" registrar ipv4:" + str(PSTN_GW_DB[1]) + " expires 3600\n")
ConfigFile.write(" transport tcp tls v1.2\n")
ConfigFile.write(" xfer target dial-peer\n")
ConfigFile.write(" crypto signaling default trustpoint " + str(PSTN_GW_DB[0]) + "\r\n") 

ConfigFile.write("voice class codec 1\n")
ConfigFile.write(" codec preference 1 g711alaw\n")
ConfigFile.write(" codec preference 2 g711ulaw\n")
ConfigFile.write(" codec preference 3 g729r8\n")
ConfigFile.write(" codec preference 4 g729br8\r\n")

ConfigFile.write("no voice class sip-profiles 200\n")
ConfigFile.write("voice class sip-profiles 200\n")
ConfigFile.write(" response 200 sip-header Remote-Party-ID remove\n")
ConfigFile.write(" response 183 sip-header Remote-Party-ID remove\r\n") 
########################################################################################
############################## Dial-Plan ###############################################
ConfigFile.write("####### Dial-PLAN #######\n\r")
ConfigFile.write("no voice class e164-pattern-map 1\n")
ConfigFile.write("voice class e164-pattern-map 1\n")
ConfigFile.write(" description ***Pattern to Cluster***\n")
ConfigFile.write("e164 +" + str(DailPlanDB[0]) + str(DailPlanDB[1]) + str(DailPlanDB[2]) + "\r\n")

ConfigFile.write("voice translation-rule 10\n")
ConfigFile.write(" rule 1 /\(^.+\)/ /+49\\1/ type national national plan any isdn\n")
ConfigFile.write(" rule 2 /\(^.+\)/ /+\\1/ type international international plan any isdn\r\n")

ConfigFile.write("voice translation-rule 20\n")
ConfigFile.write(" rule 1 /^" + str(PSTNno49_REGEX[2:]) + str(DailPlanDB[1]) + "0/ /+" + str(DailPlanDB[0]) + str(DailPlanDB[1])+"999/\n")
ConfigFile.write(" rule 2 /^" + str(PSTNno49_REGEX[2:]) + str(DailPlanDB[1]) + "/ /+49\\0/\n")
ConfigFile.write(" rule 3 /^" + str(PSTNno49_REGEX[3:]) + str(DailPlanDB[1]) + "/ /+" + str(DailPlanDB[0]) + str(DailPlanDB[1])+"\\1/\n")
ConfigFile.write(" rule 4 /^" + str(PSTNno49_REGEX[4:]) + str(DailPlanDB[1]) + "/ /+" + str(DailPlanDB[0]) + str(DailPlanDB[1])+"\\1/\n")
ConfigFile.write(" rule 5 /^" + str(PSTNno49_REGEX[5:]) + str(DailPlanDB[1]) + "/ /+" + str(DailPlanDB[0]) + str(DailPlanDB[1])+"\\1/\n")
ConfigFile.write(" rule 6 /^" + str(DailPlanDB[1]) + "/ /+" + str(DailPlanDB[0]) + str(DailPlanDB[1])+"\\1/\n\r")

ConfigFile.write("voice translation-rule 30\n")
ConfigFile.write(" rule 1 /\+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"999/ /0/ type any national plan any isdn\n")
ConfigFile.write(" rule 2 /\+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) + "\(...\)/ /\\1/ type any national plan any isdn\n")
ConfigFile.write(" rule 4 /^$/ /0/\n\r")

ConfigFile.write("voice translation-rule 40\n")
ConfigFile.write(" rule 5 /^00049/ // type any national plan any isdn\n")
ConfigFile.write(" rule 6 /^000/ // type any international plan any isdn\n")
ConfigFile.write(" rule 7 /^00/ // type any national plan any isdn\n")
ConfigFile.write(" rule 8 /^0/ // type any subscriber plan any isdn\n")
ConfigFile.write(" rule 9 /^\+49/ // type any national plan any isdn\n")
ConfigFile.write(" rule 10 /^\+/ // type any international plan any isdn\n\r")

ConfigFile.write("voice translation-rule 50\n")
ConfigFile.write(" rule 1 /^\([1-9]..\)/ /+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"\\1/\n")
ConfigFile.write(" rule 2 /^000/ /+/\n")
ConfigFile.write(" rule 3 /^00/ /+49/\n")
ConfigFile.write(" rule 4 /^0\([2-9]\)/ /+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"\\1/\n")
ConfigFile.write(" rule 6 /\+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"\(0\)/ /+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"888/\n")
ConfigFile.write(" rule 7 /\+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"\(999\)/ /+"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"888/\n\r")

ConfigFile.write("voice translation-profile PSTN-INCOMING\n")
ConfigFile.write(" translate calling 10\n")
ConfigFile.write(" translate called 20\n\r")

ConfigFile.write("voice translation-profile PSTN-OUTGOING\n")
ConfigFile.write(" translate calling 30\n")
ConfigFile.write(" translate called 40\n\r")

ConfigFile.write("voice translation-profile SRST-INCOMING\n")
ConfigFile.write(" translate called 50\n\r")

###################################################################################################
##############################SCCP Dynamic Config #################################################
ConfigFile.write("####### SCCP Dynamic Config #######\n\r")

ConfigFile.write("no voice class server-group 1\n")
ConfigFile.write("voice class server-group 1\n")
ConfigFile.write("ipv4 " + str(CUCM_Group[5]) + "\n")
ConfigFile.write("ipv4 " + str(CUCM_Group[6]) +  " preference 1\n")
ConfigFile.write("description " + str(CUCM_Group[0]) +" ***"  + str(CUCM_Group[3]) +" - "  + str(CUCM_Group[4]) +"***\n\r")

ConfigFile.write("no voice class uri ClusterIPs sip\n")
ConfigFile.write("voice class uri ClusterIPs sip\n")
ConfigFile.write(" host ipv4:10.216.1.6\n")
ConfigFile.write(" host ipv4:10.216.3.14\n")
ConfigFile.write(" host ipv4:10.216.3.15\n")
ConfigFile.write(" host ipv4:10.216.1.7\n")
ConfigFile.write(" host ipv4:10.216.3.10\n")
ConfigFile.write(" host ipv4:10.216.1.10\n")
ConfigFile.write(" host ipv4:10.216.1.11\n\r")

ConfigFile.write("sccp local Loopback0\n")
ConfigFile.write("sccp ccm " + str(CUCM_Group[5]) +" identifier 1 version 7.0\n")
ConfigFile.write("sccp ccm " + str(CUCM_Group[6]) +" identifier 2 version 7.0\n")
ConfigFile.write("sccp ccm " + str(PSTN_GW_DB[1]) +" identifier 3 version 7.0\n")
ConfigFile.write("sccp ccm group 1\n")
ConfigFile.write(" bind interface Loopback0\n")
ConfigFile.write(" associate ccm 1 priority 1\n")
ConfigFile.write(" associate ccm 2 priority 2\n")
ConfigFile.write(" associate ccm 3 priority 3\n")
ConfigFile.write(" associate profile 1 register "+ str(PSTN_GW_DB[0]) +"\n")
ConfigFile.write("ccm-manager music-on-hold bind Loopback0\n\r")

ConfigFile.write("sccp\n\r")

ConfigFile.write("dspfarm profile 1 conference security\n")
ConfigFile.write(" trustpoint "+ str(PSTN_GW_DB[0]) +"\n")
ConfigFile.write(" codec g729br8\n")
ConfigFile.write(" codec g729r8\n")
ConfigFile.write(" codec g729abr8\n")
ConfigFile.write(" codec g729ar8\n")
ConfigFile.write(" codec g711alaw\n")
ConfigFile.write(" codec g711ulaw\n")
ConfigFile.write(" maximum sessions 3\n")
ConfigFile.write(" associate application SCCP\n")
ConfigFile.write(" no shut\n\r")
########################################################################################
############################## Dial-PEERS ###############################################
ConfigFile.write("####### Dial-Peers #######\n\r")
ConfigFile.write("dial-peer voice 1000 pots\n")
ConfigFile.write(" description ***Inbound POTS Main Number***\n")
ConfigFile.write(" translation-profile incoming PSTN-INCOMING\n")
ConfigFile.write(" incoming called-number "+ str(PSTNno49_REGEX[2:]) + str(DailPlanDB[1]) +"[0]\n")
ConfigFile.write(" direct-inward-dial\n\r")

ConfigFile.write("dial-peer voice 1001 pots\n")
ConfigFile.write(" description ***Inbound POTS Number Range***\n")
ConfigFile.write(" translation-profile incoming PSTN-INCOMING\n")
ConfigFile.write(" incoming called-number "+ str(PSTNno49_REGEX[2:]) + str(DailPlanDB[1]) + str(DailPlanDB[2]) +"\n")
ConfigFile.write(" direct-inward-dial\n\r")

ConfigFile.write("dial-peer voice 1100 voip\n")
ConfigFile.write(" description ***To and From CUCM Cluster***\n")
ConfigFile.write(" session protocol sipv2\n")
ConfigFile.write(" session transport tcp tls\n")
ConfigFile.write(" session server-group 1\n")
ConfigFile.write(" destination e164-pattern-map 1\n")
ConfigFile.write(" incoming uri via ClusterIPs\n")
ConfigFile.write(" voice-class codec 1\n")
ConfigFile.write(" voice-class sip profiles 200\n")
ConfigFile.write(" voice-class sip options-keepalive\n")
ConfigFile.write(" dtmf-relay rtp-nte\n")
ConfigFile.write(" fax-relay ecm disable\n")
ConfigFile.write(" no vad\n\r")

ConfigFile.write("dial-peer voice 1200 pots\n")
ConfigFile.write(" trunkgroup PSTN\n")
ConfigFile.write(" description ***Outgoing POTS enbloc***\n")
ConfigFile.write(" translation-profile outgoing PSTN-OUTGOING\n")
ConfigFile.write(" destination-pattern 0T\n")
ConfigFile.write(" progress_ind alert enable 8\n")
ConfigFile.write(" progress_ind progress enable 8\n\r")

ConfigFile.write("dial-peer voice 1300 pots\n")
ConfigFile.write(" trunkgroup PSTN\n")
ConfigFile.write(" description ***Outgoing POTS E164-Dialing***\n")
ConfigFile.write(" translation-profile outgoing PSTN-OUTGOING\n")
ConfigFile.write(" destination-pattern +T\n")
ConfigFile.write(" progress_ind alert enable 8\n")
ConfigFile.write(" progress_ind progress enable 8\n")
ConfigFile.write(" progress_ind connect enable 8\n\r")

ConfigFile.write("do show dial-peer voice 1100 | i state\n\r")

########################################################################################
##############################SIP SRST #################################################
ConfigFile.write("####### SIP SRST #######\n\r")
ConfigFile.write("voice register global\n")
ConfigFile.write(" default mode\n")
ConfigFile.write(" no allow-hash-in-dn\n")
ConfigFile.write(" security-policy secure\n")
ConfigFile.write(" system message Lokale Notfalltelefonie\n")
ConfigFile.write(" max-dn 300\n")
ConfigFile.write(" max-pool 100\n\r")

ConfigFile.write("no voice register pool  1\n")
ConfigFile.write("voice register pool  1\n")
ConfigFile.write(" translation-profile incoming SRST-INCOMING\n")
ConfigFile.write(" translation-profile outgoing SRST-INCOMING\n")
ConfigFile.write(" id network 10.0.0.0 mask 255.0.0.0\n")
ConfigFile.write(" alias 1 +"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"0 to +"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"888\n")
ConfigFile.write(" alias 2 +"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"999 to +"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +"888\n")
ConfigFile.write(" dialplan-pattern 1 +"+ str(DailPlanDB[0]) + str(DailPlanDB[1]) +str(DailPlanDB[2]) +" extension-length 3\n")
ConfigFile.write(" no digit collect kpml\n")
ConfigFile.write(" presence call-list\n")
ConfigFile.write(" dtmf-relay rtp-nte\n")
ConfigFile.write(" codec g711alaw\n")
ConfigFile.write(" no vad\n\r")

###################################################################################################
##############################ISDN Interfaces Dynamic Config #################################################
ConfigFile.write("####### ISDN Dynamic Config #######\n\r")
if SearchISDN[0]=="PRI":
    if SearchISDN[1]==1:
       ConfigFile.write("isdn switch-type primary-net5\n\r")
       ConfigFile.write("card type e1 0 1\n\r")
       ConfigFile.write("voice-card 0/1\n")
       ConfigFile.write(" no watchdog\n\r")

       ConfigFile.write("voice-port 0/1/0:15\n")
       ConfigFile.write("sh\n")
       ConfigFile.write("echo-cancel enable\n")
       ConfigFile.write("compand-type a-law\n")
       ConfigFile.write("cptone DE\n")
       ConfigFile.write("bearer-cap 3100Hz\n\r")

       ConfigFile.write("controller E1 0/1/0\n")
       ConfigFile.write(" pri-group timeslots 1-31\n")
       ConfigFile.write(" clock source line primary\n\r")
       ConfigFile.write("interface Serial0/1/0:15\n")
       ConfigFile.write(" shut\n")
       ConfigFile.write(" no ip address\n")
       ConfigFile.write(" encapsulation hdlc\n")
       ConfigFile.write(" isdn switch-type primary-net5\n")
       ConfigFile.write(" isdn overlap-receiving\n")
       ConfigFile.write(" isdn incoming-voice voice\n")
       ConfigFile.write(" isdn outgoing display-ie\n")
       ConfigFile.write(" trunk-group PSTN\n")
       ConfigFile.write(" no shut\n\r")
       ConfigFile.write("voice-port 0/1/0:15\n")
       ConfigFile.write("no sh\n")

    elif SearchISDN[1]==2:
       ConfigFile.write("isdn switch-type primary-net5\n\r")
       ConfigFile.write("card type e1 0 1\n\r")
       ConfigFile.write("card type e1 0 2\n\r")
       ConfigFile.write("voice-card 0/1\n")
       ConfigFile.write(" no watchdog\n\r")
       ConfigFile.write("voice-card 0/2\n")
       ConfigFile.write(" no watchdog\n\r")

       ConfigFile.write("voice-port 0/1/0:15\n")
       ConfigFile.write("sh\n")
       ConfigFile.write("echo-cancel enable\n")
       ConfigFile.write("compand-type a-law\n")
       ConfigFile.write("cptone DE\n")
       ConfigFile.write("bearer-cap 3100Hz\n\r")

       ConfigFile.write("voice-port 0/2/0:15\n")
       ConfigFile.write("sh\n")
       ConfigFile.write("echo-cancel enable\n")
       ConfigFile.write("compand-type a-law\n")
       ConfigFile.write("cptone DE\n")
       ConfigFile.write("bearer-cap 3100Hz\n\r")

       ConfigFile.write("controller E1 0/1/0\n")
       ConfigFile.write(" pri-group timeslots 1-31\n")
       ConfigFile.write(" clock source line primary\n\r")
       ConfigFile.write("interface Serial0/1/0:15\n")
       ConfigFile.write(" shut\n")
       ConfigFile.write(" no ip address\n")
       ConfigFile.write(" encapsulation hdlc\n")
       ConfigFile.write(" isdn switch-type primary-net5\n")
       ConfigFile.write(" isdn overlap-receiving\n")
       ConfigFile.write(" isdn incoming-voice voice\n")
       ConfigFile.write(" isdn outgoing display-ie\n")
       ConfigFile.write(" trunk-group PSTN\n")
       ConfigFile.write(" no shut\n\r")

       ConfigFile.write("controller E1 0/2/0\n")
       ConfigFile.write(" pri-group timeslots 1-31\n")
       ConfigFile.write(" clock source line primary\n\r")
       ConfigFile.write("interface Serial0/2/0:15\n")
       ConfigFile.write(" shut\n")
       ConfigFile.write(" no ip address\n")
       ConfigFile.write(" encapsulation hdlc\n")
       ConfigFile.write(" isdn switch-type primary-net5\n")
       ConfigFile.write(" isdn overlap-receiving\n")
       ConfigFile.write(" isdn incoming-voice voice\n")
       ConfigFile.write(" isdn outgoing display-ie\n")
       ConfigFile.write(" trunk-group PSTN\n")
       ConfigFile.write(" no shut\n\r")

       ConfigFile.write("voice-port 0/1/0:15\n")
       ConfigFile.write("no sh\n\r")

       ConfigFile.write("voice-port 0/2/0:15\n")
       ConfigFile.write("no sh\n\r")

elif SearchISDN[0]=="BRI":
    if SearchISDN[1]==1:
        ConfigFile.write("isdn switch-type basic-net3\n\r")
        ConfigFile.write("interface BRI0/1/0:0\n")
        ConfigFile.write(" sh\n")
        ConfigFile.write(" no ip address\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" isdn overlap-receiving\n")
        ConfigFile.write(" isdn point-to-point-setup\n")
        ConfigFile.write(" isdn incoming-voice voice\n")
        ConfigFile.write(" isdn static-tei 0\n")
        ConfigFile.write(" trunk-group PSTN\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" no sh\n\r")

        ConfigFile.write("interface BRI0/1/1:0\n")
        ConfigFile.write(" sh\n")
        ConfigFile.write(" no ip address\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" isdn overlap-receiving\n")
        ConfigFile.write(" isdn point-to-point-setup\n")
        ConfigFile.write(" isdn incoming-voice voice\n")
        ConfigFile.write(" isdn static-tei 0\n")
        ConfigFile.write(" trunk-group PSTN\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" no sh\n\r")

        ConfigFile.write("voice-port 0/1/0\n")
        ConfigFile.write(" echo-cancel enable\n")
        ConfigFile.write(" compand-type a-law\n")
        ConfigFile.write(" cptone DE\n")
        ConfigFile.write(" bearer-cap 3100Hz\n\r")
    
        ConfigFile.write("voice-port 0/1/1\n")
        ConfigFile.write(" echo-cancel enable\n")
        ConfigFile.write(" compand-type a-law\n")
        ConfigFile.write(" cptone DE\n")
        ConfigFile.write(" bearer-cap 3100Hz\n\r")


    elif SearchISDN[1]==2:
        ConfigFile.write("isdn switch-type basic-net3\n\r")

        ConfigFile.write("interface BRI0/1/0:0\n")
        ConfigFile.write(" sh\n")
        ConfigFile.write(" no ip address\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" isdn overlap-receiving\n")
        ConfigFile.write(" isdn point-to-point-setup\n")
        ConfigFile.write(" isdn incoming-voice voice\n")
        ConfigFile.write(" isdn static-tei 0\n")
        ConfigFile.write(" trunk-group PSTN\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" no sh\n\r")

        ConfigFile.write("interface BRI0/1/1:0\n")
        ConfigFile.write(" sh\n")
        ConfigFile.write(" no ip address\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" isdn overlap-receiving\n")
        ConfigFile.write(" isdn point-to-point-setup\n")
        ConfigFile.write(" isdn incoming-voice voice\n")
        ConfigFile.write(" isdn static-tei 0\n")
        ConfigFile.write(" trunk-group PSTN\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" no sh\n\r")

        ConfigFile.write("interface BRI0/2/0:0\n")
        ConfigFile.write(" sh\n")
        ConfigFile.write(" no ip address\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" isdn overlap-receiving\n")
        ConfigFile.write(" isdn point-to-point-setup\n")
        ConfigFile.write(" isdn incoming-voice voice\n")
        ConfigFile.write(" isdn static-tei 0\n")
        ConfigFile.write(" trunk-group PSTN\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" no sh\n\r")

        ConfigFile.write("interface BRI0/2/1:0\n")
        ConfigFile.write(" sh\n")
        ConfigFile.write(" no ip address\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" isdn overlap-receiving\n")
        ConfigFile.write(" isdn point-to-point-setup\n")
        ConfigFile.write(" isdn incoming-voice voice\n")
        ConfigFile.write(" isdn static-tei 0\n")
        ConfigFile.write(" trunk-group PSTN\n")
        ConfigFile.write(" isdn switch-type basic-net3\n")
        ConfigFile.write(" no sh\n\r")

        ConfigFile.write("voice-port 0/1/0\n")
        ConfigFile.write(" echo-cancel enable\n")
        ConfigFile.write(" compand-type a-law\n")
        ConfigFile.write(" cptone DE\n")
        ConfigFile.write(" bearer-cap 3100Hz\n\r")

        ConfigFile.write("voice-port 0/1/1\n")
        ConfigFile.write(" echo-cancel enable\n")
        ConfigFile.write(" compand-type a-law\n")
        ConfigFile.write(" cptone DE\n")
        ConfigFile.write(" bearer-cap 3100Hz\n\r")

        ConfigFile.write("voice-port 0/2/0\n")
        ConfigFile.write(" echo-cancel enable\n")
        ConfigFile.write(" compand-type a-law\n")
        ConfigFile.write(" cptone DE\n")
        ConfigFile.write(" bearer-cap 3100Hz\n\r")

        ConfigFile.write("voice-port 0/2/1\n")
        ConfigFile.write(" echo-cancel enable\n")
        ConfigFile.write(" compand-type a-law\n")
        ConfigFile.write(" cptone DE\n")
        ConfigFile.write(" bearer-cap 3100Hz\n\r")
else:
    print("ISDN data is not avallable")
    ConfigFile.write("####### ISDN Dynamic Config is not avallable #######\n\r")

ConfigFile.close()
print("config file is generated:"+str(ConfigFile)+" in location:"+str(py_file_location)+"\\"+str(SiteCode))
###################################################################################################
############################## Analog GW config #################################################
###################### Create Config File ########################################################
ConfigFile= open(str(SiteCode) + "_ANALOG_GW_1_config.txt","w+")
ConfigFile.write("########### Config File for Analog GW 1 " + str(AnalogGW1_DB[0]) + " ########### \r\n")
ConfigFile.close()
ConfigFile= open(str(SiteCode) + "_ANALOG_GW_1_config.txt","a+")
############################## Analog GW-1 config #################################################
ConfigFile.write("####### Analog GW-1 config #######" + str(AnalogGW1_DB[0]) + "\n\r")
ConfigFile.write("####### Analog GW Certificate #######\n\r")

ConfigFile.write("crypto key generate rsa general-keys label " + str(AnalogGW1_DB[0]) + " modulus 2048\n\r")

ConfigFile.write("do show interfaces gigabitEthernet 0/1 | i bia\n\r")

ConfigFile.write("crypto pki trustpoint " + str(AnalogGW1_DB[0]) + "\n")
ConfigFile.write(" enrollment terminal pem\n")
ConfigFile.write(" serial-number none\n")
ConfigFile.write(" fqdn none\n")
ConfigFile.write(" ip-address none\n")
ConfigFile.write(" subject-name CN=84:a1:33:60:69 (LAST 10 !!! ) GigabitEthernet0/1  bia 84:a1:33:60:69)\n")
ConfigFile.write(" revocation-check none\n")
ConfigFile.write(" rsakeypair " + str(AnalogGW1_DB[0]) + " 2048 2048\n\r")

ConfigFile.write("crypto pki enroll " + str(AnalogGW1_DB[0]) + "\n\r")

ConfigFile.write("Install first the ROOT cert\n")
ConfigFile.write("crypto pki authenticate " + str(AnalogGW1_DB[0]) + "\n\r")

ConfigFile.write("Install second the CA\n")
ConfigFile.write("crypto pki import " + str(AnalogGW1_DB[0]) + " certificate\n\r")

############################## Analog GW-1 VoIP config #################################################
ConfigFile.write("####### Analog GW VoIP #######\n\r")

ConfigFile.write("no stcapp\n")
ConfigFile.write("stcapp ccm-group 1\n")
ConfigFile.write("stcapp security trustpoint "+ str(AnalogGW1_DB[0]) +"\n")
ConfigFile.write("stcapp security tls-version v1.2\n")
ConfigFile.write("stcapp security mode encrypted\n")
ConfigFile.write("stcapp\n\r")

ConfigFile.write("no sccp\n")
ConfigFile.write("no sccp ccm group 1\n")
ConfigFile.write("no sccp ccm 10.216.3.10 identifier 2 version 7.0\n")
ConfigFile.write("no sccp ccm 10.216.1.10 identifier 1 version 7.0\n")
ConfigFile.write("no sccp ccm "+ str(PSTN_GW_DB[1]) + " identifier 3 version 7.0\n")
ConfigFile.write("do show run | i sccp ccm 10.\n\r")

ConfigFile.write("sccp local GigabitEthernet0/1\n")
ConfigFile.write("sccp ccm " + str(CUCM_Group[5]) +" identifier 1 version 7.0\n")
ConfigFile.write("sccp ccm " + str(CUCM_Group[6]) +" identifier 2 version 7.0\n")
ConfigFile.write("sccp ccm " + str(PSTN_GW_DB[1]) +" identifier 3 version 7.0\n")
ConfigFile.write("sccp\n\r")

ConfigFile.write("sccp ccm group 1\n")
ConfigFile.write(" bind interface GigabitEthernet0/1\n")
ConfigFile.write(" associate ccm 1 priority 1\n")
ConfigFile.write(" associate ccm 2 priority 2\n")
ConfigFile.write(" associate ccm 3 priority 3\n\r")

ConfigFile.write("ccm-manager config server 10.216.3.7 10.216.1.12\n")
ConfigFile.write("ccm-manager sccp local GigabitEthernet0/1\n")
ConfigFile.write("ccm-manager sccp\n\r")

ConfigFile.write("snmp-server view UC_SNMP dod included\n")
ConfigFile.write("snmp-server host 10.216.3.17 version 3 auth UC_SNMP\n")
ConfigFile.write("snmp-server group UC_SNMP v3 auth read UC_SNMP write UC_SNMP\n")
ConfigFile.write("snmp-server user UC_SNMP UC_SNMP v3 auth sha UC_SNMP2019!\n\r")

ConfigFile.close()
print("config file is generated:"+str(ConfigFile)+" in location:"+str(py_file_location)+"\\"+str(SiteCode))
